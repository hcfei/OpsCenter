# -*- coding: utf-8 -*-
"""读取预测结果模板.xlsx 的完整结构与数据"""
import openpyxl, json, sys

PATH = r"D:\工作\03. 领域运营\02.运营预测\预测结果模板.xlsx"

wb = openpyxl.load_workbook(PATH, data_only=False)  # 保留公式
print("=== 文件信息 ===")
print("Sheet 列表:", wb.sheetnames)
print()

for ws in wb.worksheets:
    print("=" * 70)
    print(f"### Sheet: {ws.title} | 维度: {ws.max_row}行 x {ws.max_column}列 | 状态: {ws.sheet_state}")
    print("=" * 70)
    # 合并单元格
    if ws.merged_cells.ranges:
        merges = [str(r) for r in ws.merged_cells.ranges]
        print("合并单元格:", merges[:30], "..." if len(merges) > 30 else "")
    # 读取所有单元格
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=min(ws.max_column, 20)), start=1):
        cells = []
        for c in row:
            if c.value is not None:
                cells.append(f"{c.coordinate}={repr(c.value)}")
        if cells:
            print(f"  R{row_idx}: " + " | ".join(cells))
    print()
