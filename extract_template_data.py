# -*- coding: utf-8 -*-
"""提取预测结果模板数据为结构化 JSON"""
import openpyxl, json

PATH = r"D:\工作\03. 领域运营\02.运营预测\预测结果模板.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb['预算目标拆分']

# 表头定义: 列 -> 周期字段
ACT = {'C':'q1_act','D':'q2_act','E':'q3_act','F':'q4_act','G':'h1_act','H':'h2_act','I':'ytd_act'}
TGT = {'J':'q1_tgt','K':'q2_tgt','L':'q3_tgt','M':'q4_tgt','N':'h1_tgt','O':'h2_tgt','P':'ytd_tgt'}
# 偏差/达成率由前端计算, 不入库

# 处理 A 列合并单元格向下填充
def fill_merged(col_letter):
    vals = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1 and mr.max_col == 1:  # A 列合并
            top = ws[f'{col_letter}{mr.min_row}'].value
            for rr in range(mr.min_row, mr.max_row + 1):
                vals[rr] = top
    return vals

metric_fill = fill_merged('A')

rows = []
for r in range(3, 18):
    metric = metric_fill.get(r) or ws.cell(row=r, column=1).value  # A列 指标 (合并填充)
    bu = ws.cell(row=r, column=2).value       # B列 BU
    rec = {'metric': metric, 'bu': bu}
    ok = False
    for col, f in ACT.items():
        v = ws[f'{col}{r}'].value
        rec[f] = round(float(v), 6) if isinstance(v, (int, float)) else v
        ok = ok or (v is not None)
    for col, f in TGT.items():
        v = ws[f'{col}{r}'].value
        rec[f] = round(float(v), 6) if isinstance(v, (int, float)) else v
    if ok and metric and bu:
        rows.append(rec)

print(f"提取 {len(rows)} 条记录")
for rec in rows:
    print(json.dumps(rec, ensure_ascii=False))

# 汇总校验
print("\n=== 校验: 汇总行 = 各BU之和 (以收入YTD实际为例) ===")
bu_sum = sum(r['ytd_act'] for r in rows if r['bu'] != '汇总' and r['metric'] == '收入')
total = [r for r in rows if r['bu'] == '汇总' and r['metric'] == '收入'][0]['ytd_act']
print(f"BU合计: {bu_sum:.6f} vs 汇总行: {total:.6f}")

with open('forecast_template_data.json', 'w', encoding='utf-8') as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)
print("\n已保存 forecast_template_data.json")
